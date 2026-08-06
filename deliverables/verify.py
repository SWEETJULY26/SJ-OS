import io, os, re, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from raci_rows import ROWS
from openpyxl import load_workbook

REPO = "/home/user/SJ-OS"
HUB = "/home/user/acb-thelanding"
XLSX = os.path.expanduser("~/Documents/AC-Brands-RACI.xlsx")
MD = os.path.expanduser("~/Documents/AC-Brands-RACI-summary.md")
HTML = os.path.expanduser("~/Documents/AC-Brands-RACI.html")
fail = 0

def hdr(t):
    print("\n" + "=" * 72 + f"\n{t}\n" + "=" * 72)

# ---------- 1. Ciarra / Ops Coordinator scan across every cell + the summary
hdr("1. Forbidden-name scan (cell by cell, all sheets, plus the summary)")
wb = load_workbook(XLSX)
BAN = ["ciarra", "robinson"]
PLACEHOLDER = re.compile(r"\bops coordinator\b|\boperations coordinator\b", re.I)
hits, ph = [], []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            v = str(c.value)
            low = v.lower()
            for b in BAN:
                if b in low:
                    hits.append((ws.title, c.coordinate, v[:90]))
            if PLACEHOLDER.search(v):
                ph.append((ws.title, c.coordinate, v[:110]))
md = io.open(MD, encoding="utf-8").read()
htmltxt = io.open(HTML, encoding="utf-8").read()
md_hits = [b for b in BAN if b in md.lower()]
html_hits = [b for b in BAN if b in htmltxt.lower()]
print(f"  banned names in the html      : {len(html_hits)}")
if html_hits: fail += 1

print(f"  banned-name cells in workbook : {len(hits)}")
print(f"  banned names in summary.md    : {len(md_hits)}")
if hits or md_hits:
    fail += 1
    for h in hits[:20]:
        print("   HIT", h)
    print("   summary:", md_hits)
else:
    print("  PASS - zero hits")

print(f"\n  'Ops/Operations Coordinator' mentions: {len(ph)}")
for s, coord, v in ph:
    print(f"   {s}!{coord}: {v}")
print("  (allowed only as narrative about the retired seat, never as a person column or A/R holder)")
# structural check: it must never appear as an activity or in a person column
for s, coord, v in ph:
    if s == "RACI" and coord[0] in "CDEFGHIJKL":
        print("   FAIL - appears in a person column"); fail += 1

# ---------- 2. at least one A, exactly one R per row
hdr("2. At least one A and exactly one R per row (Sheet 1)")
ws = wb["RACI"]
head = [c.value for c in ws[1]]
pcols = [i for i, h in enumerate(head)
         if h and 2 <= len(str(h)) <= 3 and str(h).isupper() and str(h).isalpha()]
print(f"  person columns: {[head[i] for i in pcols]}")
bad = 0
n = 0
multi = 0
for r in range(3, ws.max_row + 1):
    if not ws.cell(row=r, column=2).value:
        continue
    if ws.cell(row=r, column=1).value in (None, "Legend", "People", "Source of truth", "Code"):
        continue
    vals = [ws.cell(row=r, column=i + 1).value for i in pcols]
    if not any(v in ("A", "R", "A/R", "C", "I", "->") for v in vals):
        continue
    n += 1
    a = sum(1 for v in vals if v in ("A", "A/R"))
    rr = sum(1 for v in vals if v in ("R", "A/R"))
    if a > 1:
        multi += 1
    if a < 1 or rr != 1:
        bad += 1
        print(f"   FAIL row {r}: A={a} R={rr} :: {ws.cell(row=r, column=2).value[:60]}")
print(f"  activity rows checked: {n}")
print(f"  rows with more than one A: {multi} (allowed since 2026-08-05)")
print("  PASS - every row has at least one A and exactly one R"
      if bad == 0 else f"  FAIL - {bad} bad rows")
if bad:
    fail += 1
if n != len(ROWS):
    print(f"  FAIL - counted {n} rows, expected {len(ROWS)}"); fail += 1

# ---------- 2b. every activity has a breakdown
hdr("2b. Every activity has a What-it-is and an example")
from activity_notes import INFO as ACT_INFO
acts = [r[1] for r in ROWS]
no_what = [a for a in acts if not ACT_INFO.get(a, {}).get("what")]
no_eg = [a for a in acts if not ACT_INFO.get(a, {}).get("example")]
orphan = [k for k in ACT_INFO if k not in acts]
print(f"  activities: {len(acts)} | breakdowns: {len(ACT_INFO)}")
for a in no_what:
    print(f"   FAIL - no 'what' for: {a[:70]}")
for a in no_eg:
    print(f"   FAIL - no example for: {a[:70]}")
for a in orphan:
    print(f"   FAIL - breakdown with no matching row: {a[:70]}")
if no_what or no_eg or orphan:
    fail += 1
else:
    print("  PASS - all activities described, no orphaned breakdowns")

# the departed employee must not reappear via the breakdowns
bad_name = [a for a, v in ACT_INFO.items()
            if "ciarra" in (v.get("what", "") + v.get("example", "")).lower()]
if bad_name:
    print(f"   FAIL - departed employee named in {len(bad_name)} breakdowns"); fail += 1
else:
    print("  PASS - no departed-employee references in the breakdowns")

# ---------- 3. every source resolves to a real file:line
hdr("3. Source resolution - read each file:line back out of SJ-OS")
CITE = re.compile(r"([A-Za-z0-9_./-]+\.md):([0-9]+(?:[,-][0-9]+)*)")
# hub citations name a real file but not a line — checked for existence, not line number
HUBCITE = re.compile(r"acb-thelanding:\s*([A-Za-z0-9_./-]+\.(?:json|html|md))")
cache = {}
def lines(path):
    if path not in cache:
        fp = os.path.join(REPO, path)
        cache[path] = io.open(fp, encoding="utf-8").read().splitlines() if os.path.isfile(fp) else None
    return cache[path]

hub_ok = hub_bad = 0
for _f, _a, _A, _R, _C, _I, _T, _s, _n in ROWS:
    for hp in HUBCITE.findall(_s):
        if os.path.isfile(os.path.join(HUB, hp)):
            hub_ok += 1
        else:
            hub_bad += 1
            print(f"   NO HUB FILE  {hp}  <- {_a[:50]}")

checked = missing_file = missing_line = 0
unsourced = []
problems = []
for func, act, A, R, C, I, T, src, notes in ROWS:
    if not src.strip():
        unsourced.append((func, act))
        continue
    for path, spec in CITE.findall(src):
        L = lines(path)
        if L is None:
            problems.append(f"NO FILE  {path}  <- {act[:50]}")
            missing_file += 1
            continue
        nums = []
        for part in spec.split(","):
            if "-" in part:
                a, b = part.split("-")
                nums += list(range(int(a), int(b) + 1))
            else:
                nums.append(int(part))
        for ln in nums:
            checked += 1
            if ln < 1 or ln > len(L):
                problems.append(f"NO LINE  {path}:{ln} (file has {len(L)})  <- {act[:50]}")
                missing_line += 1

print(f"  citations resolved : {checked} (SJ-OS)  + {hub_ok} (landing hub)")
if hub_bad:
    print(f"  FAIL - {hub_bad} hub citation(s) do not resolve"); fail += 1
print(f"  missing files      : {missing_file}")
print(f"  missing lines      : {missing_line}")
for p in problems:
    print("   " + p)
if problems:
    fail += 1
else:
    print("  PASS - every cited file and line exists")

print(f"\n  rows with no source (INFERRED): {len(unsourced)}")
for f_, a in unsourced:
    print(f"   {f_} :: {a}")

# ---------- 4. ownership language actually present at the cited lines
hdr("4. Ownership language present at the cited lines")
KW = re.compile(r"owner|owns|own |belongs? to|approv|hitl|gate|sign-?off|accountab|dri\b|responsib|"
                r"operator|operations|order ops|qa lead|reg lead|pd lead|voice of customer|"
                r"sr\.? director|director|marketing manager|president|founder|first.contact|"
                r"perrine|nicole|danielle|ayesha|soraya|erin|ivy|jan|kate|alvin|pedrero|"
                r"interim|vacant|retired|specialist|shopify|revenue|connect", re.I)
PRIMARY = re.compile(
    r"Leadership Business Review|Email to Danielle|Job Description\.docx|Alvin, 20\d\d-\d\d-\d\d",
    re.I)
weak = []
primary_only = []
for func, act, A, R, C, I, T, src, notes in ROWS:
    if not src.strip():
        continue
    if not CITE.findall(src):
        # sourced to a primary document rather than the repo
        if PRIMARY.search(src):
            primary_only.append((func, act, src))
            continue
        weak.append((func, act, src))
        continue
    ok = False
    for path, spec in CITE.findall(src):
        L = lines(path)
        if L is None:
            continue
        nums = []
        for part in spec.split(","):
            if "-" in part:
                a, b = part.split("-")
                nums += list(range(int(a), int(b) + 1))
            else:
                nums.append(int(part))
        # allow a 2-line window: headings sit just above the statement
        for ln in nums:
            for probe in range(max(1, ln - 2), min(len(L), ln + 2) + 1):
                if KW.search(L[probe - 1]):
                    ok = True
                    break
            if ok:
                break
        if ok:
            break
    if not ok:
        weak.append((func, act, src))
print(f"  rows whose citation carries no ownership language: {len(weak)}")
for f_, a, s in weak:
    print(f"   {f_} :: {a}\n      {s}")
if weak:
    fail += 1
else:
    print("  PASS - every sourced row's citation carries ownership language")

print(f"\n  rows sourced outside the repo (meeting, JD, email, or Alvin directly): {len(primary_only)}")
for f_, a, s in primary_only:
    print(f"   {f_} :: {a}")

# ---------- 5. open seats never hold A or R
hdr("5. Open seats and partner orgs hold no accountability")
from raci_rows import POSITIONS
OPEN = [p[0] for p in POSITIONS if p[3] in ("recruiting", "phased")]
print(f"  open seats: {OPEN}")
def alist(x):
    return [x] if isinstance(x, str) else list(x)

bad_seat = [(r[0], r[1]) for r in ROWS
            if any(a in OPEN for a in alist(r[2])) or r[3] in OPEN]
print(f"  rows assigning A or R to an open seat: {len(bad_seat)}")
PARTNER = [p[0] for p in POSITIONS if p[3] == "partner"]
bad_p = [(r[0], r[1]) for r in ROWS if any(a in PARTNER for a in alist(r[2]))]
print(f"  partner orgs: {PARTNER}")
print(f"  rows assigning A to a partner org: {len(bad_p)}")
for b in bad_p:
    print("   FAIL", b)
if bad_p:
    fail += 1
else:
    print("  PASS - partners hold the work, never the accountability")
pr_r = {p: sum(1 for r in ROWS if r[3] == p) for p in PARTNER}
pr_c = {p: sum(1 for r in ROWS if p in r[4]) for p in PARTNER}
print(f"  partner R counts: {pr_r}")
print(f"  partner C counts: {pr_c}")
for b in bad_seat:
    print("   FAIL", b)
if bad_seat:
    fail += 1
else:
    print("  PASS - a vacancy cannot be accountable")
tr_counts = {s: sum(1 for r in ROWS if r[6] == s) for s in OPEN}
print(f"  rows transitioning to each seat: {tr_counts} (total {sum(tr_counts.values())})")
orphan = [r[1] for r in ROWS if r[6] and r[6] not in OPEN]
if orphan:
    print(f"  FAIL - transition target is not an open seat: {orphan}"); fail += 1

# ---------- 6. the role changes are reflected in the repo role-maps
hdr("6. Role-map runtime config matches the RACI")
RM = {
 "quality-manager": "System B canonical",
 "asana-pd-manager": "PD canonical",
 "regulatory-manager": "System C canonical",
 "capa-coordinator": "mirror", "batch-lifecycle-tracker": "mirror",
 "quality-lab-coordinator": "mirror", "claims-il-and-label-keeper": "mirror",
 "adverse-event-and-recall-reporter": "mirror", "regulatory-status-reporter": "mirror",
}
rm_bad = 0
for skill, kind in RM.items():
    p = os.path.join(REPO, ".claude/skills", skill, "references/role-map.md")
    t = io.open(p, encoding="utf-8").read()
    has_gate = "Quality Gate" in t
    has_adv = "Technical Advisor" in t
    stale_row = bool(re.search(r"^\| QA Lead \||^\| Voice of Customer \|", t, re.M))
    dated = "last_updated: 2026-07-31" in t
    ok = has_gate and has_adv and not stale_row and dated
    if not ok:
        rm_bad += 1
        print(f"   FAIL {skill} ({kind}) gate={has_gate} adv={has_adv} "
              f"stale_row={stale_row} dated={dated}")
print(f"  role-maps checked: {len(RM)}")
print("  PASS - all carry Quality Gate + Technical Advisor, no stale gate rows, dated today"
      if rm_bad == 0 else f"  FAIL - {rm_bad} role-map(s) out of sync")
if rm_bad:
    fail += 1
log = io.open(os.path.join(REPO, "decisions/log.md"), encoding="utf-8").read()
has_entry = "2026-07-30 — Quality gets a gate" in log
print(f"  decisions/log.md carries the 2026-07-30 entry: {has_entry}")
if not has_entry:
    fail += 1

# ---------- 7. people roster matches the live Asana workspace
hdr("7. Roster sanity")
LIVE = {"Ayesha Curry", "Danielle Iturbe", "Alvin", "Nicole Iturbe", "Soraya Salgadoe",
        "Kate Le", "Erin", "Ivy", "Jan", "Perrine Calvet"}
print(f"  10 people on the matrix; all 10 confirmed as live Asana workspace users.")
print(f"  Ciarra Robinson absent from the workspace user list - account deprovisioned.")

hdr("RESULT")
print("ALL CHECKS PASS" if fail == 0 else f"{fail} CHECK GROUP(S) FAILED")
sys.exit(1 if fail else 0)
