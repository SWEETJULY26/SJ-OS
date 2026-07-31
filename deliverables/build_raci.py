import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from raci_rows import ROWS, POSITIONS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.expanduser("~/Documents/AC-Brands-RACI.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

INI = [p[0] for p in POSITIONS]
NAME = {p[0]: p[1] for p in POSITIONS}
TITLE = {p[0]: p[2] for p in POSITIONS}
STATUS = {p[0]: p[3] for p in POSITIONS}
FUNCS = []
for r in ROWS:
    if r[0] not in FUNCS:
        FUNCS.append(r[0])

def counts(i):
    return dict(
        a=sum(1 for r in ROWS if r[2] == i),
        r=sum(1 for r in ROWS if r[3] == i),
        ar=sum(1 for r in ROWS if r[2] == i == r[3]),
        out=sum(1 for r in ROWS if r[3] == i and r[6]),
        inc=sum(1 for r in ROWS if r[6] == i))
CNT = {i: counts(i) for i in INI}

A = "Arial"
F_B = Font(name=A, size=10)
F_BD = Font(name=A, size=10, bold=True)
F_H = Font(name=A, size=10, bold=True, color="FFFFFF")
F_SUB = Font(name=A, size=8, color="FFFFFF")
F_T = Font(name=A, size=13, bold=True)
F_AR = Font(name=A, size=10, bold=True, color="FFFFFF")
FILL_H = PatternFill("solid", fgColor="1F3B4D")
FILL_A = PatternFill("solid", fgColor="8A665A")   # Pava Brown — accountable
FILL_R = PatternFill("solid", fgColor="6E8AAE")   # Coffee Fix, darkened — responsible
FILL_FN = PatternFill("solid", fgColor="EDF1F5")
FILL_OPEN = PatternFill("solid", fgColor="F3D54E")  # Pineapple Punch — open seat
THIN = Side(style="thin", color="D0D7DE")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")

wb = Workbook()

# ================================================================ Sheet 1
ws = wb.active
ws.title = "RACI"
NCOL = 2 + len(INI) + 3   # Function, Activity, positions, Transitions to, Source, Notes

# two-row header: position code, then holder
r1 = ["Function", "Activity"] + INI + ["Transitions to", "Source", "Notes"]
r2 = ["", ""] + [NAME[i] if STATUS[i] != "recruiting" and STATUS[i] != "phased" else "OPEN"
                 for i in INI] + ["", "", ""]
ws.append(r1)
ws.append(r2)
for c in range(1, NCOL + 1):
    a1, a2 = ws.cell(row=1, column=c), ws.cell(row=2, column=c)
    a1.font, a1.fill, a1.border = F_H, FILL_H, BD
    a2.font, a2.fill, a2.border = F_SUB, FILL_H, BD
    a1.alignment = CTR if 3 <= c <= 2 + len(INI) else Alignment(vertical="center", wrap_text=True)
    a2.alignment = CTR
for j, i in enumerate(INI):
    if STATUS[i] in ("recruiting", "phased"):
        ws.cell(row=2, column=3 + j).fill = FILL_OPEN
        ws.cell(row=2, column=3 + j).font = Font(name=A, size=8, bold=True, color="4A3C14")

for f, act, Aa, Rr, C, I, T, src, notes in ROWS:
    letters = {}
    letters[Aa] = "A/R" if Aa == Rr else "A"
    if Rr != Aa:
        letters[Rr] = "R"
    for k in C:
        letters.setdefault(k, "C")
    for k in I:
        letters.setdefault(k, "I")
    if T:
        letters.setdefault(T, "->")
    ws.append([f, act] + [letters.get(i, "") for i in INI]
              + [TITLE[T] if T else "", src, notes])

last = ws.max_row
for r in range(3, last + 1):
    for c in range(1, NCOL + 1):
        cell = ws.cell(row=r, column=c)
        cell.font, cell.border = F_B, BD
        if 3 <= c <= 2 + len(INI):
            cell.alignment = CTR
            v = cell.value
            if v in ("A", "A/R"):
                cell.fill, cell.font = FILL_A, F_AR
            elif v == "R":
                cell.fill, cell.font = FILL_R, F_AR
        else:
            cell.alignment = WRAP
    ws.cell(row=r, column=1).fill = FILL_FN
    ws.cell(row=r, column=1).font = F_BD

ws.freeze_panes = "C3"
ws.auto_filter.ref = f"A2:{get_column_letter(NCOL)}{last}"
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 62
for j in range(len(INI)):
    ws.column_dimensions[get_column_letter(3 + j)].width = 6
ws.column_dimensions[get_column_letter(3 + len(INI))].width = 26
ws.column_dimensions[get_column_letter(4 + len(INI))].width = 58
ws.column_dimensions[get_column_letter(5 + len(INI))].width = 82
ws.row_dimensions[1].height = 28

# legend + roster
lr = last + 2
ws.cell(row=lr, column=1, value="Legend").font = F_T
for j, (txt, fill) in enumerate([
    ("A = answers for the outcome (exactly one per row)", FILL_A),
    ("R = does the work (exactly one per row)", FILL_R),
    ("A/R = the same person holds both, so the row has no second pair of eyes", FILL_A),
    ("C = consulted before the decision", None),
    ("I = informed after it", None),
    ("-> = the open seat that absorbs R on hire. Open seats never hold A or R.", None),
]):
    ws.cell(row=lr + 1 + j, column=2, value=txt).font = F_B
    if fill:
        ws.cell(row=lr + 1 + j, column=1).fill = fill

pr = lr + 8
ws.cell(row=pr, column=1, value="Positions").font = F_T
for k, h in enumerate(["Code", "Position", "Holder", "Status", "A count", "R count",
                        "Both", "Hands off", "Absorbs"]):
    cell = ws.cell(row=pr + 1, column=1 + k, value=h)
    cell.font = F_BD
for j, i in enumerate(INI):
    c = CNT[i]
    row = pr + 2 + j
    vals = [i, TITLE[i], NAME[i], STATUS[i], c["a"], c["r"], c["ar"], c["out"], c["inc"]]
    for k, v in enumerate(vals):
        cell = ws.cell(row=row, column=1 + k, value=v)
        cell.font = F_B
        if k >= 4:
            cell.alignment = CTR
    if STATUS[i] in ("recruiting", "phased"):
        ws.cell(row=row, column=4).fill = FILL_OPEN

nr = pr + 2 + len(INI) + 1
ws.cell(row=nr, column=1, value="Source of truth").font = F_BD
ws.cell(row=nr, column=2, value=(
    "Rows citing a path are relative to the SJ-OS repo. Rows prefixed acb-thelanding cite the AC "
    "Brands landing hub, whose links.json defines eleven functions with a named lead each. Rows "
    "citing the 30 July leadership review, the 27 July email to Danielle, a job description, or "
    "Alvin directly come from those sources. Every row on this matrix has an owner - there are no "
    "inferred owners left.")).font = F_B
ws.cell(row=nr, column=2).alignment = WRAP

# ================================================================ Sheet 2
an = wb.create_sheet("Analysis")
for col, w in zip("ABCDEFGH", [112, 9, 9, 9, 11, 10, 10, 10]):
    an.column_dimensions[col].width = w

def para(text, bold=False, size=10):
    r = an.max_row + 1 if an["A1"].value else 1
    cell = an.cell(row=r, column=1, value=text)
    cell.font = Font(name=A, size=size, bold=bold)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    an.row_dimensions[r].height = max(15, 13 * (len(text) // 100 + 1))
    an.cell(row=r + 1, column=1, value="")
    return r

para("Ownership concentration, coverage gaps, and what the two seats move", bold=True, size=13)
para("Counts are live formulas over the RACI sheet, so they move if a row changes. "
     "Read against the matrix. Reflects the 30 July 2026 leadership review and the 31 July "
     "role corrections.")

hr = an.max_row + 1
for k, h in enumerate(["Position and holder", "A", "R", "A/R", "Hands off", "Absorbs", "A after", "R after"]):
    cell = an.cell(row=hr, column=1 + k, value=h)
    cell.font, cell.fill = F_H, FILL_H
    cell.alignment = CTR if k else Alignment(vertical="center")
nrows = len(ROWS) + 2
for j, i in enumerate(INI):
    r = hr + 1 + j
    col = get_column_letter(3 + j)
    rng = f"RACI!{col}3:{col}{nrows}"
    an.cell(row=r, column=1, value=f"{i} — {TITLE[i]} ({NAME[i]})").font = F_B
    an.cell(row=r, column=2, value=f'=COUNTIF({rng},"A")+COUNTIF({rng},"A/R")').font = F_B
    an.cell(row=r, column=3, value=f'=COUNTIF({rng},"R")+COUNTIF({rng},"A/R")').font = F_B
    an.cell(row=r, column=4, value=f'=COUNTIF({rng},"A/R")').font = F_B
    an.cell(row=r, column=5, value=CNT[i]["out"]).font = F_B
    an.cell(row=r, column=6, value=CNT[i]["inc"]).font = F_B
    an.cell(row=r, column=7, value=f"=B{r}").font = F_B
    an.cell(row=r, column=8, value=f"=C{r}-E{r}+F{r}").font = F_B
    for cc in range(2, 9):
        an.cell(row=r, column=cc).alignment = CTR
tr = hr + 1 + len(INI)
an.cell(row=tr, column=1, value="Total activities").font = F_BD
an.cell(row=tr, column=2, value=f"=COUNTA(RACI!B3:B{nrows})").font = F_BD
an.cell(row=tr, column=2).alignment = CTR
an.cell(row=tr + 1, column=1, value="")
an.cell(row=tr + 2, column=1, value=(
    '"A after" and "R after" project the book once both specialist seats are filled. '
    "Hands off counts rows where this position does the work today and an incoming seat takes it. "
    "Absorbs counts rows arriving on hire.")).font = Font(name=A, size=9, italic=True)
an.cell(row=tr + 2, column=1).alignment = WRAP
an.cell(row=tr + 3, column=1, value="")

BLOCKS = [
 ("Quality now has a gate, and it is not the VP", True),
 ("The 30 July leadership review moved the quality function onto Nicole. Verbatim from the recording: "
  "\"Nicole taking the primary focus on the quality management side of things with quality control... "
  "anything related to quality of service or quality of product will go through Nicole's eyes as sort "
  "of that final quality check... Even with each function having or being the primary owner, we'll "
  "still have that quality gate.\" That shows up as 10 of the 13 Quality rows accountable to Nicole, "
  "against 2 for Alvin and 1 for Perrine. Her total book of accountability goes from 8 rows to 21.", False),

 ("Perrine moves to technical guidance", True),
 ("She keeps A on the 7 rows where the judgment is genuinely formulation - formula stage-gate, "
  "compatibility and stability and RIPT and PET, in-market stability testing, reformulation, "
  "PD-linked receipt, and the two margin rows that turn on formulation cost. Packaging development "
  "moved to Erin Hover as the lead technical authority on packaging and artwork. She moves "
  "to consulted on everything process-shaped: NCR and CAPA lifecycle, batch hold and release, vendor "
  "quality flags, lab finding classification, SOP ratification. Danielle's framing in the review was "
  "that there were things the team had assumed Perrine would manage or organize that she is not, and "
  "that leadership can pull the plug even when she has submitted an approval. Splitting technical "
  "judgment from process ownership is what that means in practice.", False),

 ("Three contractors hold accountability, and two of them hold it alone", True),
 ("Erin Hover is the lead technical authority on packaging and artwork and is accountable for 4 rows "
  "- creative direction, packaging development, artwork execution, and the label artwork archive. "
  "She is also a contractor, and on creative direction she is both A and R with no internal "
  "counterpart. Jan Haeck, who executes under her, is a contractor too, so the whole packaging and "
  "artwork chain runs through people outside the company. Perrine Calvet is the third, accountable "
  "for 7 formulation rows. Between them that is 11 of 79 activities where nobody on staff can check "
  "the work or continue it. Neither specialist job description covers creative or formulation "
  "authority, so the two hires do not close this.", False),

 ("Nobody is blank on a product decision", True),
 ("Perrine and Soraya now hold a letter on all eight Product Development rows. Perrine was blank on "
  "one - founder brand-line approval - and Soraya was blank on all eight, which meant the marketing "
  "read had no formal visibility into any product decision. Both are Informed at minimum now. That "
  "is a small change with a real effect: a blank cell on a RACI is ambiguous in a way that Informed "
  "is not, because blank does not tell you whether the omission was deliberate.", False),

 ("Marketing and the channels finally have owners with real activities under them", True),
 ("Soraya goes from 1 accountable row to 8. Marketing now carries editorial calendar, Klaviyo email, "
  "paid media, social content, influencer and earned media, the WITHIN relationship, product copy and "
  "PDP content - none of which existed on the matrix before, because none of it has a skill behind "
  "it. Nicole goes from 21 to 26 with all-channel accountability plus wholesale pipeline, retail "
  "price architecture and the retailer promo calendar. Retail & Wholesale went from 2 rows to 5, "
  "Marketing from 9 to 15. A skill suite is a good source for what is automated; it is a poor "
  "source for what a business does.", False),

 ("Where the President and the Founder sit", True),
 ("Danielle Iturbe is accountable for 8 rows - direction-changing PD decisions, the margin walk-away "
  "call, brand guideline custody, campaign direction, month-end close, the annual budget, and the two "
  "web rows - and is consulted on 25 more. She now "
  "appears on all 9 Marketing rows and all 3 Creative rows, which she did not before. Campaign "
  "direction was missing from the matrix entirely; adding it is what gave the President somewhere to "
  "sit in Marketing rather than only receiving finished work. Ayesha Curry keeps sole accountability "
  "for brand-line moves and is consulted on 6 rows where the brand carries her name - brand "
  "guideline custody, creative direction, campaign direction, claim substantiation, retailer "
  "attestations, and packaging. Informed on the rest.", False),

 ("Alvin owns the framework, and still does most of the work", True),
 ("Accountable for 46 of 102 activities and doing the work on 58. The framework rows are deliberate - "
  "quality management system, SOP framework, this RACI, the skill suite, the 23 scheduled Routines. "
  "The problem is not the A column. It is that 31 rows have him as both A and R, and that he does the "
  "work on 11 of the 12 Operations rows and 5 of the 8 Product Development rows. He holds the gates "
  "and does the work behind them.", False),

 ("What the two seats actually move", True),
 ("32 rows transition on hire: 16 to the Operations Specialist and 16 to the Product Development "
  "Specialist. 29 of those come off Alvin, 3 off Nicole. His R book drops from 55 to 26. Rows with no "
  "second pair of eyes drop from 42 to 25 across the whole matrix. Both seats report to Nicole, which "
  "is the point Danielle made in the review - moving day-to-day management off the VP seat is the "
  "succession outcome the system is being built for.", False),

 ("Ops first is the right order, and the matrix says why", True),
 ("Alvin does the work on 11 of 12 Operations rows. The Operations Specialist job description covers "
  "almost exactly that list: daily order operations and the third-party logistics relationship, "
  "channel operations across the direct store and Ulta Beauty Marketplace and Amazon, inventory "
  "reconciliation and FEFO and out-of-stock signals, receiving, purchase-to-pay, freight and customs "
  "and broker relationships, routing-guide and advance-ship-notice compliance, sales and operations "
  "planning support, production scheduling, and the label and seeding and sampling projects. That is "
  "a defined function sitting on one person today.", False),

 ("The PD seat is the one that closes the quality loop", True),
 ("Its job description names running the quality system outright - corrective and preventive actions, "
  "non-conformances, complaint intake and trend monitoring, lab results, supplier quality flags, batch "
  "lifecycle including stability and hold-release, and the quality dashboard - plus pre-launch "
  "ingredient and label review, claim substantiation, retailer compliance responses and the "
  "registration tracker, and document control across specifications and dielines and artwork versions "
  "and bills of materials. Until it is filled, Nicole gates quality and Alvin executes it. That works, "
  "and it is the reason the gate is worth having now rather than after the hire, but it is two people "
  "covering a role that was written as one.", False),

 ("Nothing on this matrix is unowned any more", True),
 ("The first pass reported three unowned functions. All three are now owned, and none of them were "
  "ever actually ownerless - they were missing from the sources. Finance and HR needed the partner "
  "organisations added. Marketing, wholesale and Shopify revenue needed the AC Brands landing hub, "
  "which defines eleven functions with an explicit lead per function and had three reading \"Owned "
  "by TBD\": Marketing, Finance, and the wholesale half of Sales & Commerce. Resolving those took "
  "three decisions rather than three hires. Soraya is accountable for Marketing. Nicole is "
  "accountable for all channels - DTC, Amazon and wholesale together, which is broader than the hub "
  "had it. Danielle is accountable for web with Nicole as the systems and tech owner. That is 102 "
  "activities across eleven functions with an owner on every row.", False),

 ("Six functions run entirely outside the company", True),
 ("Regulatory goes through Pedrero, Finance through Ironclad (Dan Bender), HR through Calm HR as PEO "
  "and co-employer, digital marketing and email through WITHIN, web development through Teknologics. "
  "Add the three contractors on packaging, creative and formulation and that is eight external "
  "parties holding the execution of eight functions. For a team of nine it is a reasonable shape - "
  "it is how this surface gets covered at all. What it means is that continuity rests on eight "
  "engagements rather than on employment, and each is a renewal decision rather than a retention "
  "one. Engagement terms are recorded for Pedrero only.", False),

 ("Regulatory is the odd one, and it cuts both ways", True),
 ("Pedrero is consulted on 11 rows and responsible for none. That is not an undervaluation - our own "
  "procedures make them consult-only with no Asana access and no internal authority, and Amy Pedrero "
  "holds the binding regulatory calls. The consequence is a split bottleneck: Pedrero forms every "
  "regulatory opinion, and Alvin performs every regulatory action, because the Operator and Reg Lead "
  "gates are both his and Pedrero cannot act inside our systems. Either the Reg Lead gate gets an "
  "alternate or Pedrero's remit widens; the current shape means regulatory work stops if either side "
  "is unavailable.", False),

 ("One row is still unowned", True),
 ("Shopify revenue and channel position. It is registered as the revenue connection and named as the "
  "primary direct-to-consumer revenue track, but no process assigns ownership or an approval step, so "
  "it shows Alvin by inference rather than by source. The function closest to the top line is the "
  "least specified.", False),

 ("Single points of failure that remain after both hires", True),
 ("Perrine is an external contractor holding A on 8 rows with no internal counterpart, and one of "
  "those - compatibility and stability and RIPT and PET - has her as both A and R. Nobody inside AC "
  "Brands can check that work or continue it. Alvin is the other: the Operator gate has no alternate "
  "anywhere in the suite, and all 23 scheduled Routines stop at a human approval only he can clear. "
  "Succession is documented exactly once in the entire system, for competitive intelligence. Nothing "
  "comparable exists for the quality gate, the technical advisor, or the framework owner. That gap "
  "costs hours to close, not headcount, and it is the one I would fix first.", False),
]
for text, bold in BLOCKS:
    para(text, bold=bold, size=11 if bold else 10)

# ================================================================ Sheet 3
gp = wb.create_sheet("Gaps")
gh = ["Type", "Function", "Gap", "Why it matters", "Owner / next step", "Est. hrs / month", "Source"]
gp.append(gh)
for c in range(1, len(gh) + 1):
    cell = gp.cell(row=1, column=c)
    cell.font, cell.fill, cell.border = F_H, FILL_H, BD
    cell.alignment = Alignment(vertical="center", wrap_text=True)

GAPS = [
 ("Open seat", "Operations & Supply Chain",
  "Operations Specialist. Approved, prioritized, recruiting now. Reports to Nicole. $72-90K.",
  "Alvin does the work on 11 of 12 Operations rows. This seat absorbs 16 activities across order ops, "
  "inventory, purchase-to-pay, freight and planning.",
  "Recruiting via Calm HR once the RACI clears Danielle", 160,
  "decisions/log.md:52; Leadership Business Review 2026-07-30 52:48-53:17; Ops Specialist JD"),

 ("Open seat", "Product Development",
  "PD Specialist. Approved, phased in after Ops. Reports to Nicole. $75-95K.",
  "Absorbs 16 activities including the whole quality-execution layer the JD calls 'run the quality "
  "system', plus document control and product regulatory work. Until then Nicole gates and Alvin executes.",
  "Define scope, then recruit after Ops seat lands", 160,
  "decisions/log.md:52; Leadership Business Review 2026-07-30 53:17-57:10; PD Specialist JD"),

 ("Resolved", "Finance",
  "Finance is owned. Ironclad Finance (Dan Bender) runs AP, bookkeeping, payroll, month-end close "
  "and reporting. Danielle accountable for reporting, Alvin for cost.",
  "Was flagged as unowned before the external partners were added to the matrix. The remaining "
  "question is not ownership but coverage: no internal finance headcount exists, so continuity "
  "depends on the Ironclad engagement.",
  "No action - confirm engagement terms and renewal date", 0,
  "Alvin, 2026-07-31 (external partners)"),

 ("Resolved", "People & Admin",
  "HR is owned. Calm HR is the PEO and co-employer and runs handbook, policy, benefits, payroll "
  "admin, onboarding, offboarding and separations. Alvin accountable as liaison, Danielle co-approves.",
  "Was flagged as unowned before the external partners were added. What remains internal is the "
  "system-side half of offboarding - Asana reassignment, wiki contact state, access deprovisioning - "
  "which the departed-role-holder checklist covers and which is still on Alvin.",
  "No action on HR; system-side offboarding to the Ops Specialist", 4,
  "Alvin, 2026-07-31 (external partners); asana-pd-manager/references/role-map.md:48-56"),

 ("Resolved", "Ecommerce & DTC",
  "Shopify revenue and channel position is owned. Nicole is accountable for all channels - DTC, "
  "Amazon and wholesale - with Alvin doing the work and Danielle informed.",
  "Was the last row on the matrix with no owner. The landing hub had split DTC and Amazon to Alvin "
  "and left wholesale as TBD; consolidating all channels under one accountability closes it.",
  "No action - every row on the matrix now has an owner", 0,
  "acb-thelanding: data/links.json; Alvin, 2026-07-31 (marketing, wholesale and web)"),

 ("Resolved", "Marketing & Brand",
  "Marketing is owned. Soraya is accountable across editorial, email, paid media and social. "
  "WITHIN responsible for paid and email execution, Kate for social content.",
  "The landing hub listed Marketing as \"Owned by TBD\" with four named tools and nobody against "
  "them. Fifteen Marketing rows now have an owner.",
  "No action - confirm WITHIN engagement terms and renewal", 0,
  "acb-thelanding: data/links.json; Alvin, 2026-07-31 (marketing, wholesale and web)"),

 ("Single point of failure", "Product Development",
  "Compatibility, stability, RIPT and PET decisions sit with an external contractor as both A and R.",
  "No internal counterpart can check the work or continue it. Pre-launch technical decisions have no "
  "second reader.",
  "PD Specialist as internal counterpart on hire", 0,
  "asana-pd-manager/references/role-map.md:16"),

 ("Concentration", "All functions",
  "Six whole functions run through external partner organisations: Regulatory (Pedrero), Finance "
  "(Ironclad), HR (Calm HR), digital marketing (WITHIN), web development (Teknologics), managed IT "
  "(Coastal Interactive). Plus three contractors on packaging, creative and formulation.",
  "Nine external parties hold the execution of Regulatory, Finance, HR, digital marketing, web "
  "development, managed IT, packaging, creative and formulation. Not a fault - it is how a team of "
  "nine covers this surface - but continuity rests on nine engagements rather than on employment, "
  "and each one is a renewal decision. Engagement terms are recorded for Pedrero only.",
  "Record engagement terms and renewal dates for all nine", 0,
  "Alvin, 2026-07-31 (external partners)"),

 ("Single point of failure", "Regulatory & Compliance",
  "Pedrero forms every regulatory opinion but holds no accountability, and Alvin performs every "
  "regulatory action. Consulted on 11 rows, responsible for none.",
  "Our own procedures make Pedrero consult-only with no Asana access and no internal authority, so "
  "the binding external call and the internal execution are both bottlenecks in different ways: "
  "their opinion gates the work, and only Alvin can act on it.",
  "Give the Reg Lead gate an alternate, or widen Pedrero's remit", 0,
  "regulatory-manager/references/pedrero-contacts.md:9,13,21"),

 ("Single point of failure", "Creative",
  "Erin Hover is a contractor and the lead technical authority on packaging and artwork, "
  "accountable for 4 rows with no internal backup. On creative direction she is both A and R.",
  "Jan Haeck executes under her and is also a contractor, so the entire packaging and artwork "
  "chain sits outside the company. Neither approved job description covers creative authority.",
  "Name an internal creative counterpart, or scope it into a future role", 0,
  "Alvin, 2026-07-31 (role correction); asana-pd-manager/references/role-map.md:22-23"),

 ("Single point of failure", "Quality",
  "Perrine holds A on 7 formulation rows as a contractor, with no named backup.",
  "Reduced from 11 by moving process gates to Nicole and packaging to Erin, which is real progress. "
  "What remains is genuinely formulation and genuinely single-threaded.",
  "Name a backup technical reviewer", 0,
  "Leadership Business Review 2026-07-30 25:18-25:50; quality-manager/references/role-map.md:14"),

 ("Single point of failure", "Regulatory & Compliance",
  "Operator and Reg Lead are the same person, so two gates designed to be independent fire against one approver.",
  "On ingredient-list packets, retailer attestations and FDA filings the second approval adds an audit "
  "entry but no second judgment. The 15-day reporting clock has no alternate.",
  "Split Reg Lead, or route the second gate to Pedrero", 0,
  "regulatory-manager/references/role-map.md:14; adverse-event-and-recall-reporter/SKILL.md:108"),

 ("Single point of failure", "IT / Systems & Data",
  "All 23 scheduled Routines stop at a human approval only Alvin can clear.",
  "Automation that fires unattended still needs him to commit. A week away queues the work rather than "
  "pausing it, and the Routines keep firing.",
  "Delegate publish gates to the new specialists", 0,
  "scheduled-prompts/weekly-pd-update.md:91"),

 ("Single point of failure", "All functions",
  "Succession is documented once in the whole system - Soraya as interim CI lead if Nicole is out.",
  "Nothing comparable exists for the quality gate, the technical advisor, the Reg Lead or Creative. "
  "The repo's stated purpose is succession.",
  "Alvin to write per-role backups into each role-map", 4,
  "sjs-comp-intel/references/team-ownership.md:95; decisions/log.md:41"),

 ("Open item", "Quality",
  "Quality management system and monthly trend review not yet stood up. Target end of Q3.",
  "The gate is assigned but the reporting cadence behind it is not built. Without trend data a recurring "
  "issue like the pump defect cannot drive a packaging decision.",
  "Alvin owns framework, Nicole runs it", 12,
  "Leadership Business Review 2026-07-30 26:02-26:17"),

 ("Open item", "Operations & Supply Chain",
  "22 tasks from the retired coordinator seat sit unassigned in an archived holding project; 4 overdue.",
  "An unassigned task drops out of every per-person sweep and goes stale silently. Concrete cost of "
  "running the interim split for a quarter.",
  "Alvin to reassign under the interim split", 3,
  "decisions/log.md:60"),

 ("Open item", "People & Admin",
  "SOP cleanup and operational prep before onboarding is not done.",
  "Danielle's stated condition on the hire: clean up the system first so the role can function. "
  "Onboarding into the current state recreates the too-broad seat.",
  "Alvin and Nicole, before the Ops seat starts", 16,
  "Leadership Business Review 2026-07-30 54:24-55:43, 1:00:32"),

 ("Open item", "IT / Systems & Data",
  "Wiki write-back has not fired since 2026-05-26. 123 of 133 pages are untouched seed content.",
  "The bridges' runtime lexicon reads stale seed data, degrading vendor, contact and product "
  "recognition on every sweep.",
  "Alvin; part of the skills-architecture backlog", 8,
  "decisions/wiki-layer-audit-2026-07-29.md:52; decisions/log.md:75"),

 ("Open item", "Quality",
  "Nine SOPs and forms are 29 days overdue on annual review.",
  "Review dates were null so the annual sweep had nothing to fire on. Dates are restored; the reviews "
  "are not done, and each needs sign-off.",
  "Nicole approves as the quality gate; Alvin stages", 12,
  "decisions/log.md:69"),

 ("Uncovered brand", "All functions",
  "Sweet July, the lifestyle brand, has no coverage. The suite is Sweet July Skin plus company-wide work.",
  "Merch, licensing and non-skincare retail are run somewhere, but nothing documents who owns them. "
  "They cannot be put on a RACI from these sources.",
  "Danielle to confirm scope for the next version", 0,
  "references/architecture/system_map.md:3; context/about-business.md:3"),
]
for g in GAPS:
    gp.append(list(g))

glast = gp.max_row
for r in range(2, glast + 1):
    for c in range(1, len(gh) + 1):
        cell = gp.cell(row=r, column=c)
        cell.font, cell.border = F_B, BD
        cell.alignment = CTR if c == 6 else WRAP
    gp.cell(row=r, column=1).font = F_BD
    if gp.cell(row=r, column=1).value == "Open seat":
        gp.cell(row=r, column=1).fill = FILL_OPEN

gp.freeze_panes = "A2"
gp.auto_filter.ref = f"A1:{get_column_letter(len(gh))}{glast}"
for col, w in zip("ABCDEFG", [20, 26, 52, 58, 34, 11, 50]):
    gp.column_dimensions[col].width = w
gp.row_dimensions[1].height = 28
gp.cell(row=glast + 2, column=1, value="Hours per month").font = F_BD
gp.cell(row=glast + 2, column=3, value=(
    "Rough estimates. Open-seat rows are full-time-equivalent load, not incremental. "
    "Zero means the gap is a structural exposure rather than unbilled hours.")).font = F_B
gp.cell(row=glast + 2, column=3).alignment = WRAP

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("wrote", OUT, "| rows:", len(ROWS), "| gaps:", len(GAPS))
